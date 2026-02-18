"""
Teste do endpoint /reports/pid-analysis-excel com dados reais.
"""
import json
import sys
sys.path.insert(0, ".")

from src.application.pid_excel_service import PIDExcelService
from src.domain.models import PIDAnalysisExportRequest

# JSON real do Power Automate (amostra)
raw_pages = [
    {
        "is_relevant_pid": False,
        "reasoning": "The image contains general notes/annotations about theatre subcircuits and HVAC with no visible P&ID/HVAC schematic elements."
    },
    {
        "is_relevant_pid": True,
        "reasoning": "The image is a mechanical/steam P&ID-style schematic (headers, PRV, separator SEP-01).",
        "drawing_title": "UNKNOWN",
        "drawing_number": "UNKNOWN",
        "components": [],
        "summary": {"total_devices": 0, "total_points": 0, "by_type": {"AI": 0, "AO": 0, "DI": 0, "DO": 0}}
    },
    {
        "is_relevant_pid": True,
        "reasoning": "The image shows a mechanical/hydronic schematic with existing headers and labeled heating water pumps.",
        "drawing_title": "Unknown (title block not visible)",
        "drawing_number": "Unknown",
        "components": [
            {"tag": "S1HWP-01", "device_type": "Pump", "points": ["DO (Start)", "DI (Run)", "DI (Trip)"], "point_count": 3, "confidence": "medium"},
            {"tag": "S1HWP-02", "device_type": "Pump", "points": ["DO (Start)", "DI (Run)", "DI (Trip)"], "point_count": 3, "confidence": "medium"},
            {"tag": "S1HWP-03", "device_type": "Pump", "points": ["DO (Start)", "DI (Run)", "DI (Trip)"], "point_count": 3, "confidence": "medium"}
        ],
        "summary": {"total_devices": 3, "total_points": 9, "by_type": {"AI": 0, "AO": 0, "DI": 6, "DO": 3}}
    },
    {
        "is_relevant_pid": True,
        "reasoning": "The image shows a mechanical/HVAC schematic with tagged control valve (CV-S-01).",
        "drawing_title": "Unknown",
        "drawing_number": "Unknown",
        "components": [
            {"tag": "CV-S-01", "device_type": "Control/Modulating Valve", "points": ["AO (Position)", "AI (Feedback)"], "point_count": 2, "confidence": "high"},
            {"tag": "T1", "device_type": "Temperature Transmitter", "points": ["AI (Temperature)"], "point_count": 1, "confidence": "low"}
        ],
        "summary": {"total_devices": 2, "total_points": 3, "by_type": {"AI": 2, "AO": 1, "DI": 0, "DO": 0}}
    },
    {
        "is_relevant_pid": True,
        "reasoning": "DHW equipment with clear instrumentation/control tags.",
        "drawing_title": "Not visible in provided crop",
        "drawing_number": "Unknown",
        "components": [
            {"tag": "CV-DHW-01", "device_type": "Control/Modulating Valve", "points": ["AO (Position)", "AI (Feedback)"], "point_count": 2, "confidence": "high"},
            {"tag": "CV-DHW-02", "device_type": "Control/Modulating Valve", "points": ["AO (Position)", "AI (Feedback)"], "point_count": 2, "confidence": "high"},
            {"tag": "PHP-01", "device_type": "Pump", "points": ["DO (Start)", "DI (Run)", "DI (Trip)"], "point_count": 3, "confidence": "high"},
            {"tag": "VSD-PHP-01", "device_type": "Variable Speed Drive", "points": ["DO (Enable)", "AO (Speed)", "DI (Run)", "DI (Fault)"], "point_count": 4, "confidence": "medium"}
        ],
        "summary": {"total_devices": 4, "total_points": 11, "by_type": {"AI": 2, "AO": 3, "DI": 4, "DO": 2}}
    },
    {
        "is_relevant_pid": False,
        "reasoning": "The image is a demolition layout plan without instrumentation tags or control schematics."
    }
]

data = PIDAnalysisExportRequest(
    project_name="Theatre HVAC P&ID Analysis",
    pages=raw_pages,
)

service = PIDExcelService()
buf = service.generate(data)

output_path = "test_pid_output.xlsx"
with open(output_path, "wb") as f:
    f.write(buf.read())

print(f"Excel gerado: {output_path}")
print(f"Tamanho: {buf.tell()} bytes")

# Verificar conteudo
from openpyxl import load_workbook
wb = load_workbook(output_path)
print(f"\nAbas: {wb.sheetnames}")

ws_points = wb["Points List"]
print(f"\nPoints List: {ws_points.max_row} linhas, {ws_points.max_column} colunas")
for row in ws_points.iter_rows(min_row=1, max_row=min(ws_points.max_row, 8), values_only=True):
    print(f"  {row}")

ws_summary = wb["Summary"]
print(f"\nSummary: {ws_summary.max_row} linhas")
for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, values_only=True):
    print(f"  {row}")

ws_pages = wb["All Pages"]
print(f"\nAll Pages: {ws_pages.max_row} linhas")
for row in ws_pages.iter_rows(min_row=1, max_row=ws_pages.max_row, values_only=True):
    print(f"  {row}")

print("\n✅ Teste completo!")
