import uvicorn
import io
import base64
import logging
import pdfplumber
from collections import defaultdict

from fastapi import FastAPI, HTTPException, Security, Depends
from fastapi.responses import StreamingResponse
from fastapi.security import APIKeyHeader
from pydantic import BaseModel
from typing import List, Optional

# --- Bibliotecas PDF (ReportLab) ---
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm

# --- Bibliotecas Excel (OpenPyXL) ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("BMS_API")

app = FastAPI(title="BMS: PDF & Complex Excel Generator", version="5.0.0")

API_KEY_NAME = "X-API-Key"
api_key_header = APIKeyHeader(name=API_KEY_NAME, auto_error=True)
REAL_KEY = "minha-chave-secreta-123" 

async def verify_key(key: str = Security(api_key_header)):
    if key == REAL_KEY:
        return key
    raise HTTPException(401, "Chave inválida")

# ==========================================
#           MODELOS DE DADOS
# ==========================================

class PdfRequest(BaseModel):
    arquivo_base64: str

class SectionRequest(BaseModel):
    arquivo_base64: str
    inicio_texto: str
    fim_texto: Optional[str] = None

class PointData(BaseModel):
    Descriptor: str
    Signal_Type: str
    Sensor_Hardware: Optional[str] = ""
    Notes: Optional[str] = ""

class EquipmentData(BaseModel):
    Tag: str
    Description: Optional[str] = ""
    Status: Optional[str] = ""
    Switchboard_Ref: Optional[str] = ""
    Location: Optional[str] = ""
    Points: List[PointData] = []

class SystemData(BaseModel):
    System_Name: str
    Equipment: List[EquipmentData] = []

class ProjectReportRequest(BaseModel):
    Focus_Category: Optional[str] = "General"
    Systems: List[SystemData] = []

class BMSPointData(BaseModel):
    AssetTag: str
    PointName: str
    PointType: str
    Logic: str
    IsIntegration: bool = False

class BMSPointsRequest(BaseModel):
    Points: List[BMSPointData]
    Report_Title: Optional[str] = "BMS Points List"

# ==========================================
#           SERVIÇOS (LÓGICA)
# ==========================================

def extract_text_pypdf(pdf_bytes, limit=None, maintain_layout=True):
    text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages = pdf.pages[:limit] if limit else pdf.pages
        for page in pages:
            extracted = page.extract_text(layout=maintain_layout) 
            if extracted:
                text += extracted + "\n"
    return text

def service_generate_points_excel_structured(data: ProjectReportRequest) -> io.BytesIO:
    """
    Gera Excel com células mescladas (Merged Cells) similar à imagem fornecida.
    Inclui campos: System, Tag, Description, Status, Switchboard_Ref, Location e Points Details.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Points List"

    COLOR_HEADER_MAIN = "2C3E50"
    COLOR_HEADER_SUB = "5D6D7E"
    COLOR_TEXT_WHITE = "FFFFFF"
    
    thin_border = Side(border_style="thin", color="000000")
    border_all = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    font_main_header = Font(bold=True, color=COLOR_TEXT_WHITE, size=11)
    fill_main_header = PatternFill(start_color=COLOR_HEADER_MAIN, end_color=COLOR_HEADER_MAIN, fill_type="solid")
    
    font_sub_header = Font(bold=True, color=COLOR_TEXT_WHITE, size=10)
    fill_sub_header = PatternFill(start_color=COLOR_HEADER_SUB, end_color=COLOR_HEADER_SUB, fill_type="solid")

    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Colunas principais: System, Tag, Description, Status, Switchboard_Ref, Location, Points List Details (4 colunas)
    headers = ["System", "Tag", "Description", "Status", "Switchboard Ref", "Location", "Points List Details", "", "", ""]
    ws.append(headers)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = font_main_header
        cell.fill = fill_main_header
        cell.alignment = align_top_left
        cell.border = border_all
    
    # Merge das colunas de Points Details (G1:J1)
    ws.merge_cells("G1:J1")

    current_row = 2
    systems_list = data.Systems if data.Systems else []

    for system in systems_list:
        sys_name = system.System_Name
        
        for eq in system.Equipment:
            num_points = len(eq.Points)
            rows_needed = num_points + 1 if num_points > 0 else 1
            
            end_row = current_row + rows_needed - 1

            ws.cell(row=current_row, column=1, value=sys_name).alignment = align_top_left
            ws.cell(row=current_row, column=2, value=eq.Tag).alignment = align_top_left
            ws.cell(row=current_row, column=3, value=eq.Description).alignment = align_top_left
            ws.cell(row=current_row, column=4, value=eq.Status).alignment = align_top_left
            ws.cell(row=current_row, column=5, value=eq.Switchboard_Ref).alignment = align_top_left
            ws.cell(row=current_row, column=6, value=eq.Location).alignment = align_top_left

            for r in range(current_row, end_row + 1):
                for c in range(1, 7): # Colunas A a F
                    ws.cell(row=r, column=c).border = border_all

            if rows_needed > 1:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=end_row, end_column=1) # System
                ws.merge_cells(start_row=current_row, start_column=2, end_row=end_row, end_column=2) # Tag
                ws.merge_cells(start_row=current_row, start_column=3, end_row=end_row, end_column=3) # Description
                ws.merge_cells(start_row=current_row, start_column=4, end_row=end_row, end_column=4) # Status
                ws.merge_cells(start_row=current_row, start_column=5, end_row=end_row, end_column=5) # Switchboard_Ref
                ws.merge_cells(start_row=current_row, start_column=6, end_row=end_row, end_column=6) # Location

            
            if num_points > 0:
                sub_headers = ["Descriptor", "Signal Type", "Sensor Hardware", "Notes"]
                for i, text in enumerate(sub_headers):
                    c = ws.cell(row=current_row, column=7+i, value=text)
                    c.font = font_sub_header
                    c.fill = fill_sub_header
                    c.border = border_all
                    c.alignment = align_top_left

                point_row_idx = current_row + 1
                for pt in eq.Points:
                    ws.cell(row=point_row_idx, column=7, value=pt.Descriptor).border = border_all
                    ws.cell(row=point_row_idx, column=8, value=pt.Signal_Type).border = border_all
                    ws.cell(row=point_row_idx, column=9, value=pt.Sensor_Hardware).border = border_all
                    ws.cell(row=point_row_idx, column=10, value=pt.Notes).border = border_all
                    point_row_idx += 1
            else:
                ws.cell(row=current_row, column=7, value="No Points").border = border_all
                ws.cell(row=current_row, column=8, value="-").border = border_all
                ws.cell(row=current_row, column=9, value="-").border = border_all
                ws.cell(row=current_row, column=10, value="-").border = border_all

            current_row += rows_needed

    column_widths = [15, 15, 30, 10, 18, 20, 30, 12, 20, 30]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def service_generate_bms_points_excel(data: BMSPointsRequest) -> io.BytesIO:
    """
    Gera Excel estruturado a partir de lista de pontos BMS.
    Agrupa por AssetTag e organiza em tabela com merged cells.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BMS Points List"

    # --- Estilos ---
    COLOR_HEADER_MAIN = "1F4E79"
    COLOR_HEADER_SUB = "2E75B6"
    COLOR_INTEGRATION = "FFC000"
    COLOR_AI = "C6EFCE"
    COLOR_AO = "FFEB9C"
    COLOR_DI = "BDD7EE"
    COLOR_DO = "F8CBAD"
    COLOR_TEXT_WHITE = "FFFFFF"
    
    # Bordas
    thin_border = Side(border_style="thin", color="000000")
    border_all = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Fontes e Preenchimentos
    font_main_header = Font(bold=True, color=COLOR_TEXT_WHITE, size=12)
    fill_main_header = PatternFill(start_color=COLOR_HEADER_MAIN, end_color=COLOR_HEADER_MAIN, fill_type="solid")
    
    font_sub_header = Font(bold=True, color=COLOR_TEXT_WHITE, size=10)
    fill_sub_header = PatternFill(start_color=COLOR_HEADER_SUB, end_color=COLOR_HEADER_SUB, fill_type="solid")

    fill_integration = PatternFill(start_color=COLOR_INTEGRATION, end_color=COLOR_INTEGRATION, fill_type="solid")
    fill_ai = PatternFill(start_color=COLOR_AI, end_color=COLOR_AI, fill_type="solid")
    fill_ao = PatternFill(start_color=COLOR_AO, end_color=COLOR_AO, fill_type="solid")
    fill_di = PatternFill(start_color=COLOR_DI, end_color=COLOR_DI, fill_type="solid")
    fill_do = PatternFill(start_color=COLOR_DO, end_color=COLOR_DO, fill_type="solid")

    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Asset Tag", "Point Details", "", "", ""]
    ws.append(headers)

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = font_main_header
        cell.fill = fill_main_header
        cell.alignment = align_center
        cell.border = border_all
    
    ws.merge_cells("B1:E1")

    points_by_asset = defaultdict(list)
    
    for point in data.Points:
        points_by_asset[point.AssetTag].append(point)

    current_row = 2

    for asset_tag, points in points_by_asset.items():
        num_points = len(points)
        rows_needed = num_points + 1
        end_row = current_row + rows_needed - 1

        ws.cell(row=current_row, column=1, value=asset_tag).alignment = align_top_left
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)

        for r in range(current_row, end_row + 1):
            ws.cell(row=r, column=1).border = border_all

        if rows_needed > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=end_row, end_column=1)

        sub_headers = ["Point Name", "Point Type", "Logic", "Integration"]
        for i, text in enumerate(sub_headers):
            c = ws.cell(row=current_row, column=2+i, value=text)
            c.font = font_sub_header
            c.fill = fill_sub_header
            c.border = border_all
            c.alignment = align_center

        point_row_idx = current_row + 1
        for pt in points:
            cell_name = ws.cell(row=point_row_idx, column=2, value=pt.PointName)
            cell_name.border = border_all
            cell_name.alignment = align_top_left
            
            cell_type = ws.cell(row=point_row_idx, column=3, value=pt.PointType)
            cell_type.border = border_all
            cell_type.alignment = align_center
            
            if pt.PointType == "AI":
                cell_type.fill = fill_ai
            elif pt.PointType == "AO":
                cell_type.fill = fill_ao
            elif pt.PointType == "DI":
                cell_type.fill = fill_di
            elif pt.PointType == "DO":
                cell_type.fill = fill_do
            elif "Integration" in pt.PointType:
                cell_type.fill = fill_integration
            
            cell_logic = ws.cell(row=point_row_idx, column=4, value=pt.Logic)
            cell_logic.border = border_all
            cell_logic.alignment = align_top_left
            
            cell_integration = ws.cell(row=point_row_idx, column=5, value="Yes" if pt.IsIntegration else "No")
            cell_integration.border = border_all
            cell_integration.alignment = align_center
            if pt.IsIntegration:
                cell_integration.fill = fill_integration
            
            point_row_idx += 1

        current_row += rows_needed

    column_widths = [18, 35, 18, 70, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def service_generate_pdf(data: ProjectReportRequest) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=12*mm, leftMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('TitleCustom', parent=styles['Heading1'], alignment=1, fontSize=16, spaceAfter=15)
    normal_style = ParagraphStyle('NormalCustom', parent=styles['Normal'], fontSize=9, leading=11)
    header_style = ParagraphStyle('HeaderCustom', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=colors.white)
    
    elements.append(Paragraph(f"Points List: {data.Focus_Category}", title_style))
    elements.append(Spacer(1, 5*mm))

    HEADER_COLOR = colors.Color(0.26, 0.33, 0.42)
    SUB_HEADER_COLOR = colors.Color(0.35, 0.45, 0.55)
    MAIN_COLS = [40*mm, 30*mm, 50*mm, 20*mm, 130*mm]
    SUB_COLS = [60*mm, 20*mm, 45*mm]

    main_table_data = [[Paragraph("System", header_style), Paragraph("Tag", header_style), Paragraph("Description", header_style), Paragraph("Status", header_style), Paragraph("Points Details", header_style)]]

    systems_list = data.Systems if data.Systems else []

    for system in systems_list:
        sys_name = str(system.System_Name)
        for eq in system.Equipment:
            points_data = [[Paragraph("Descriptor", header_style), Paragraph("Signal", header_style), Paragraph("Notes", header_style)]]
            if eq.Points:
                for pt in eq.Points:
                    points_data.append([Paragraph(str(pt.Descriptor), normal_style), Paragraph(str(pt.Signal_Type), normal_style), Paragraph(str(pt.Notes), normal_style)])
            else:
                points_data.append(["No points", "-", "-"])
            
            sub_table = Table(points_data, colWidths=SUB_COLS, splitByRow=1)
            sub_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), SUB_HEADER_COLOR),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            
            main_table_data.append([Paragraph(sys_name, normal_style), Paragraph(str(eq.Tag), normal_style), Paragraph(str(eq.Description), normal_style), str(eq.Status), sub_table])

    if len(main_table_data) > 1:
        main_table = Table(main_table_data, colWidths=MAIN_COLS, repeatRows=1, splitByRow=1)
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HEADER_COLOR),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        elements.append(main_table)
    else:
        elements.append(Paragraph("No data", normal_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer

# ==========================================
#           ROTAS (ENDPOINTS)
# ==========================================

@app.post("/extract-toc", dependencies=[Depends(verify_key)])
async def get_toc(req: PdfRequest):
    try:
        pdf_bytes = base64.b64decode(req.arquivo_base64)
        text = extract_text_pypdf(pdf_bytes, limit=20, maintain_layout=False)
        return {"text": text}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/extract-section", dependencies=[Depends(verify_key)])
async def get_section(req: SectionRequest):
    try:
        pdf_bytes = base64.b64decode(req.arquivo_base64)
        full_text = extract_text_pypdf(pdf_bytes, maintain_layout=True)
        idx_start = full_text.find(req.inicio_texto)
        if idx_start == -1: raise HTTPException(404, "Marcador não encontrado.")
        final_text = full_text[idx_start:]
        if req.fim_texto:
            idx_end = full_text.find(req.fim_texto, idx_start)
            if idx_end != -1: final_text = full_text[idx_start:idx_end]
        return {"section_text": final_text}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/generate-pdf", dependencies=[Depends(verify_key)])
async def generate_pdf_endpoint(data: ProjectReportRequest):
    try:
        pdf_file = service_generate_pdf(data)
        safe_name = "".join([c for c in (data.Focus_Category or "Report") if c.isalnum() or c in (' ','-','_')]).strip()
        filename = f"{safe_name}.pdf"
        return StreamingResponse(
            pdf_file, 
            media_type="application/pdf", 
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro PDF: {e}")
        raise HTTPException(500, str(e))

@app.post("/generate-excel", dependencies=[Depends(verify_key)])
async def generate_excel_endpoint(data: ProjectReportRequest):
    try:
        excel_file = service_generate_points_excel_structured(data)
        
        safe_name = "".join([c for c in (data.Focus_Category or "Report") if c.isalnum() or c in (' ','-','_')]).strip()
        filename = f"{safe_name}.xlsx"
        return StreamingResponse(
            excel_file, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro Excel: {e}")
        raise HTTPException(500, str(e))

@app.post("/generate-bms-points-excel", dependencies=[Depends(verify_key)])
async def generate_bms_points_excel_endpoint(data: BMSPointsRequest):
    """
    Gera Excel estruturado a partir de lista de pontos BMS.
    
    Aceita JSON com formato:
    {
        "Points": [
            {"AssetTag": "CCV-FCU-1-4", "PointName": "Valve Position Feedback", "PointType": "AI", "Logic": "...", "IsIntegration": false},
            ...
        ],
        "Report_Title": "BMS Points List" (opcional)
    }
    """
    try:
        excel_file = service_generate_bms_points_excel(data)
        
        safe_name = "".join([c for c in (data.Report_Title or "BMS_Points") if c.isalnum() or c in (' ','-','_')]).strip()
        filename = f"{safe_name}.xlsx"
        return StreamingResponse(
            excel_file, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro BMS Excel: {e}")
        raise HTTPException(500, str(e))

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)