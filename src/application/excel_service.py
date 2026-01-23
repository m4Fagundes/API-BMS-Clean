"""
Serviço para geração de relatórios Excel (Project Report).
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.models import ProjectReportRequest


class ExcelReportService:
    """Serviço responsável pela geração de relatórios Excel estruturados."""
    
    # Constantes de cores
    COLOR_HEADER_MAIN = "2C3E50"
    COLOR_HEADER_SUB = "5D6D7E"
    COLOR_TEXT_WHITE = "FFFFFF"
    
    # Larguras das colunas
    COLUMN_WIDTHS = [15, 15, 30, 10, 18, 20, 30, 12, 20, 30]
    
    def __init__(self):
        self._setup_styles()
    
    def _setup_styles(self):
        """Configura os estilos reutilizáveis."""
        thin_border = Side(border_style="thin", color="000000")
        self._border_all = Border(
            left=thin_border, 
            right=thin_border, 
            top=thin_border, 
            bottom=thin_border
        )
        
        self._font_main_header = Font(bold=True, color=self.COLOR_TEXT_WHITE, size=11)
        self._fill_main_header = PatternFill(
            start_color=self.COLOR_HEADER_MAIN, 
            end_color=self.COLOR_HEADER_MAIN, 
            fill_type="solid"
        )
        
        self._font_sub_header = Font(bold=True, color=self.COLOR_TEXT_WHITE, size=10)
        self._fill_sub_header = PatternFill(
            start_color=self.COLOR_HEADER_SUB, 
            end_color=self.COLOR_HEADER_SUB, 
            fill_type="solid"
        )
        
        self._align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    def generate(self, data: ProjectReportRequest) -> io.BytesIO:
        """
        Gera um Excel estruturado com células mescladas.
        
        Args:
            data: Dados do projeto para o relatório.
            
        Returns:
            Buffer contendo o arquivo Excel gerado.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Detailed Points List"
        
        self._create_header(ws)
        self._populate_data(ws, data)
        self._set_column_widths(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_header(self, ws):
        """Cria o cabeçalho da planilha."""
        headers = [
            "System", "Tag", "Description", "Status", 
            "Switchboard Ref", "Location", "Points List Details", "", "", ""
        ]
        ws.append(headers)
        
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = self._font_main_header
            cell.fill = self._fill_main_header
            cell.alignment = self._align_top_left
            cell.border = self._border_all
        
        # Merge das colunas de Points Details
        ws.merge_cells("G1:J1")
    
    def _populate_data(self, ws, data: ProjectReportRequest):
        """Popula a planilha com os dados."""
        current_row = 2
        systems_list = data.Systems if data.Systems else []
        
        for system in systems_list:
            sys_name = system.System_Name
            
            for eq in system.Equipment:
                current_row = self._add_equipment_row(
                    ws, current_row, sys_name, eq
                )
    
    def _add_equipment_row(self, ws, current_row: int, sys_name: str, eq) -> int:
        """Adiciona uma linha de equipamento com seus pontos."""
        num_points = len(eq.Points)
        rows_needed = num_points + 1 if num_points > 0 else 1
        end_row = current_row + rows_needed - 1
        
        # Dados principais do equipamento
        self._set_cell(ws, current_row, 1, sys_name)
        self._set_cell(ws, current_row, 2, eq.Tag)
        self._set_cell(ws, current_row, 3, eq.Description)
        self._set_cell(ws, current_row, 4, eq.Status)
        self._set_cell(ws, current_row, 5, eq.Switchboard_Ref)
        self._set_cell(ws, current_row, 6, eq.Location)
        
        # Bordas para colunas A-F
        for r in range(current_row, end_row + 1):
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = self._border_all
        
        # Merge cells se necessário
        if rows_needed > 1:
            self._merge_equipment_cells(ws, current_row, end_row)
        
        # Adiciona pontos ou placeholder
        if num_points > 0:
            self._add_points_subheader(ws, current_row)
            self._add_points_data(ws, current_row + 1, eq.Points)
        else:
            self._add_no_points_placeholder(ws, current_row)
        
        return current_row + rows_needed
    
    def _set_cell(self, ws, row: int, col: int, value):
        """Define valor e estilo de uma célula."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = self._align_top_left
        return cell
    
    def _merge_equipment_cells(self, ws, start_row: int, end_row: int):
        """Mescla células do equipamento."""
        for col in range(1, 7):
            ws.merge_cells(
                start_row=start_row, start_column=col, 
                end_row=end_row, end_column=col
            )
    
    def _add_points_subheader(self, ws, row: int):
        """Adiciona subcabeçalho dos pontos."""
        sub_headers = ["Descriptor", "Signal Type", "Sensor Hardware", "Notes"]
        for i, text in enumerate(sub_headers):
            cell = ws.cell(row=row, column=7+i, value=text)
            cell.font = self._font_sub_header
            cell.fill = self._fill_sub_header
            cell.border = self._border_all
            cell.alignment = self._align_top_left
    
    def _add_points_data(self, ws, start_row: int, points):
        """Adiciona dados dos pontos."""
        for idx, pt in enumerate(points):
            row = start_row + idx
            ws.cell(row=row, column=7, value=pt.Descriptor).border = self._border_all
            ws.cell(row=row, column=8, value=pt.Signal_Type).border = self._border_all
            ws.cell(row=row, column=9, value=pt.Sensor_Hardware).border = self._border_all
            ws.cell(row=row, column=10, value=pt.Notes).border = self._border_all
    
    def _add_no_points_placeholder(self, ws, row: int):
        """Adiciona placeholder quando não há pontos."""
        ws.cell(row=row, column=7, value="No Points").border = self._border_all
        ws.cell(row=row, column=8, value="-").border = self._border_all
        ws.cell(row=row, column=9, value="-").border = self._border_all
        ws.cell(row=row, column=10, value="-").border = self._border_all
    
    def _set_column_widths(self, ws):
        """Define as larguras das colunas."""
        for i, width in enumerate(self.COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
