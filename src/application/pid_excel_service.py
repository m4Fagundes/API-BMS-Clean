"""
Serviço para geração de Excel a partir de resultados de análise P&ID.
Gera duas abas: Points List (detalhada) e Summary (resumo por drawing).
"""
import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.models import PIDAnalysisExportRequest


class PIDExcelService:
    """Gera Excel estruturado a partir de análise de P&ID."""

    # Cores
    COLOR_HEADER = "1F4E79"
    COLOR_DRAWING = "2E75B6"
    COLOR_AI = "C6EFCE"
    COLOR_AO = "FFEB9C"
    COLOR_DI = "BDD7EE"
    COLOR_DO = "F8CBAD"
    COLOR_HIGH = "C6EFCE"
    COLOR_MEDIUM = "FFEB9C"
    COLOR_LOW = "F8CBAD"
    COLOR_IRRELEVANT = "D9D9D9"
    COLOR_WHITE = "FFFFFF"
    COLOR_SUMMARY_HEADER = "4472C4"

    def __init__(self):
        self._setup_styles()

    def _setup_styles(self):
        thin = Side(border_style="thin", color="000000")
        self._border = Border(left=thin, right=thin, top=thin, bottom=thin)

        self._font_header = Font(bold=True, color=self.COLOR_WHITE, size=11)
        self._fill_header = PatternFill(
            start_color=self.COLOR_HEADER,
            end_color=self.COLOR_HEADER,
            fill_type="solid",
        )
        self._font_drawing = Font(bold=True, color=self.COLOR_WHITE, size=10)
        self._fill_drawing = PatternFill(
            start_color=self.COLOR_DRAWING,
            end_color=self.COLOR_DRAWING,
            fill_type="solid",
        )
        self._fill_summary_header = PatternFill(
            start_color=self.COLOR_SUMMARY_HEADER,
            end_color=self.COLOR_SUMMARY_HEADER,
            fill_type="solid",
        )

        self._type_fills = {}
        for name, color in [
            ("AI", self.COLOR_AI),
            ("AO", self.COLOR_AO),
            ("DI", self.COLOR_DI),
            ("DO", self.COLOR_DO),
        ]:
            self._type_fills[name] = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

        self._confidence_fills = {}
        for name, color in [
            ("high", self.COLOR_HIGH),
            ("medium", self.COLOR_MEDIUM),
            ("low", self.COLOR_LOW),
        ]:
            self._confidence_fills[name] = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

        self._fill_irrelevant = PatternFill(
            start_color=self.COLOR_IRRELEVANT,
            end_color=self.COLOR_IRRELEVANT,
            fill_type="solid",
        )

        self._align_center = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self._align_left = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        self._align_top_left = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

    @staticmethod
    def _parse_point_type(point_str: str) -> str:
        """Extrai o tipo (AI/AO/DI/DO) de strings como 'AO (Position)'."""
        match = re.match(r"^(AI|AO|DI|DO)\b", point_str.strip())
        return match.group(1) if match else ""

    @staticmethod
    def _parse_point_description(point_str: str) -> str:
        """Extrai a descrição de strings como 'AO (Position)' -> 'Position'."""
        match = re.match(r"^(?:AI|AO|DI|DO)\s*\((.+)\)$", point_str.strip())
        return match.group(1).strip() if match else point_str.strip()

    # ------------------------------------------------------------------
    #  Generate
    # ------------------------------------------------------------------
    def generate(self, data: PIDAnalysisExportRequest) -> io.BytesIO:
        wb = Workbook()

        # --- Aba 1: Points List ---
        ws_points = wb.active
        ws_points.title = "Points List"
        self._build_points_sheet(ws_points, data)

        # --- Aba 2: Summary ---
        ws_summary = wb.create_sheet("Summary")
        self._build_summary_sheet(ws_summary, data)

        # --- Aba 3: All Pages ---
        ws_pages = wb.create_sheet("All Pages")
        self._build_pages_sheet(ws_pages, data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    #  Aba 1 – Points List
    # ------------------------------------------------------------------
    def _build_points_sheet(self, ws, data: PIDAnalysisExportRequest):
        headers = [
            "Page #",
            "Drawing Title",
            "Drawing Number",
            "Tag",
            "Device Type",
            "Point Type",
            "Point Description",
            "Confidence",
        ]
        widths = [8, 35, 18, 18, 28, 12, 25, 12]

        # Header row
        for col, (text, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = self._font_header
            cell.fill = self._fill_header
            cell.alignment = self._align_center
            cell.border = self._border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"
        row = 2

        for page_idx, page in enumerate(data.pages, start=1):
            if not page.is_relevant_pid:
                continue
            if not page.components:
                continue

            title = page.drawing_title or "Unknown"
            number = page.drawing_number or "Unknown"

            for comp in page.components:
                if not comp.points:
                    # Componente sem pontos detalhados
                    self._write_point_row(
                        ws, row, page_idx, title, number,
                        comp.tag, comp.device_type, "", "",
                        comp.confidence or "medium",
                    )
                    row += 1
                else:
                    for pt_str in comp.points:
                        pt_type = self._parse_point_type(pt_str)
                        pt_desc = self._parse_point_description(pt_str)
                        self._write_point_row(
                            ws, row, page_idx, title, number,
                            comp.tag, comp.device_type,
                            pt_type, pt_desc,
                            comp.confidence or "medium",
                        )
                        row += 1

    def _write_point_row(
        self, ws, row, page_idx, title, number,
        tag, device_type, pt_type, pt_desc, confidence
    ):
        values = [page_idx, title, number, tag, device_type, pt_type, pt_desc, confidence]
        aligns = [
            self._align_center,
            self._align_left,
            self._align_center,
            self._align_center,
            self._align_left,
            self._align_center,
            self._align_left,
            self._align_center,
        ]
        for col, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = self._border
            cell.alignment = align

        # Cor do tipo
        if pt_type in self._type_fills:
            ws.cell(row=row, column=6).fill = self._type_fills[pt_type]

        # Cor da confiança
        conf_lower = (confidence or "").lower()
        if conf_lower in self._confidence_fills:
            ws.cell(row=row, column=8).fill = self._confidence_fills[conf_lower]

    # ------------------------------------------------------------------
    #  Aba 2 – Summary
    # ------------------------------------------------------------------
    def _build_summary_sheet(self, ws, data: PIDAnalysisExportRequest):
        headers = [
            "Page #", "Drawing Title", "Drawing Number",
            "Devices", "Total Points", "AI", "AO", "DI", "DO",
        ]
        widths = [8, 40, 18, 10, 12, 8, 8, 8, 8]

        for col, (text, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = self._font_header
            cell.fill = self._fill_summary_header
            cell.alignment = self._align_center
            cell.border = self._border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"
        row = 2

        totals = {"devices": 0, "points": 0, "AI": 0, "AO": 0, "DI": 0, "DO": 0}

        for page_idx, page in enumerate(data.pages, start=1):
            if not page.is_relevant_pid:
                continue

            s = page.summary
            by_type = (s.by_type or {}) if s else {}
            devices = s.total_devices if s else 0
            points = s.total_points if s else 0
            ai = by_type.get("AI", 0)
            ao = by_type.get("AO", 0)
            di = by_type.get("DI", 0)
            do = by_type.get("DO", 0)

            vals = [
                page_idx,
                page.drawing_title or "Unknown",
                page.drawing_number or "Unknown",
                devices, points, ai, ao, di, do,
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = self._border
                cell.alignment = self._align_center if col != 2 else self._align_left

            # Cores AI/AO/DI/DO
            for col_offset, key in enumerate(["AI", "AO", "DI", "DO"]):
                c = ws.cell(row=row, column=6 + col_offset)
                if c.value and c.value > 0:
                    c.fill = self._type_fills[key]

            totals["devices"] += devices
            totals["points"] += points
            totals["AI"] += ai
            totals["AO"] += ao
            totals["DI"] += di
            totals["DO"] += do
            row += 1

        # Linha de totais
        total_row = row
        ws.cell(row=total_row, column=1, value="").border = self._border
        cell_total = ws.cell(row=total_row, column=2, value="TOTAL")
        cell_total.font = Font(bold=True, size=11)
        cell_total.border = self._border
        ws.cell(row=total_row, column=3, value="").border = self._border

        for col, key in [(4, "devices"), (5, "points"), (6, "AI"), (7, "AO"), (8, "DI"), (9, "DO")]:
            cell = ws.cell(row=total_row, column=col, value=totals[key])
            cell.font = Font(bold=True, size=11)
            cell.border = self._border
            cell.alignment = self._align_center

    # ------------------------------------------------------------------
    #  Aba 3 – All Pages (inclui irrelevantes)
    # ------------------------------------------------------------------
    def _build_pages_sheet(self, ws, data: PIDAnalysisExportRequest):
        headers = ["Page #", "Relevant?", "Drawing Title", "Reasoning"]
        widths = [8, 12, 40, 80]

        for col, (text, width) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = self._font_header
            cell.fill = self._fill_header
            cell.alignment = self._align_center
            cell.border = self._border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

        for page_idx, page in enumerate(data.pages, start=1):
            row = page_idx + 1
            ws.cell(row=row, column=1, value=page_idx).border = self._border
            ws.cell(row=row, column=1).alignment = self._align_center

            relevant_cell = ws.cell(
                row=row, column=2,
                value="Yes" if page.is_relevant_pid else "No",
            )
            relevant_cell.border = self._border
            relevant_cell.alignment = self._align_center
            if not page.is_relevant_pid:
                relevant_cell.fill = self._fill_irrelevant

            title_cell = ws.cell(
                row=row, column=3,
                value=page.drawing_title or ("—" if not page.is_relevant_pid else "Unknown"),
            )
            title_cell.border = self._border
            title_cell.alignment = self._align_left

            reason_cell = ws.cell(row=row, column=4, value=page.reasoning or "")
            reason_cell.border = self._border
            reason_cell.alignment = self._align_top_left
