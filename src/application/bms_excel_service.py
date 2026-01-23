"""
Serviço para geração de Excel de pontos BMS.
"""
import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.models import BMSPointsRequest


class BMSExcelService:
    """Serviço responsável pela geração de Excel de pontos BMS."""
    
    # Constantes de cores
    COLOR_HEADER_MAIN = "1F4E79"
    COLOR_HEADER_SUB = "2E75B6"
    COLOR_INTEGRATION = "FFC000"
    COLOR_AI = "C6EFCE"
    COLOR_AO = "FFEB9C"
    COLOR_DI = "BDD7EE"
    COLOR_DO = "F8CBAD"
    COLOR_TEXT_WHITE = "FFFFFF"
    
    # Larguras das colunas
    COLUMN_WIDTHS = [18, 35, 18, 70, 12]
    
    def __init__(self):
        self._setup_styles()
    
    def _setup_styles(self):
        """Configura os estilos reutilizáveis."""
        thin_border = Side(border_style="thin", color="000000")
        self._border_all = Border(
            left=thin_border, 
            right=thin_border, 
            top=thin_border, 
            bottom=thin_border
        )
        
        self._font_main_header = Font(bold=True, color=self.COLOR_TEXT_WHITE, size=12)
        self._fill_main_header = PatternFill(
            start_color=self.COLOR_HEADER_MAIN, 
            end_color=self.COLOR_HEADER_MAIN, 
            fill_type="solid"
        )
        
        self._font_sub_header = Font(bold=True, color=self.COLOR_TEXT_WHITE, size=10)
        self._fill_sub_header = PatternFill(
            start_color=self.COLOR_HEADER_SUB, 
            end_color=self.COLOR_HEADER_SUB, 
            fill_type="solid"
        )
        
        # Fills para tipos de pontos
        self._fill_integration = PatternFill(
            start_color=self.COLOR_INTEGRATION, 
            end_color=self.COLOR_INTEGRATION, 
            fill_type="solid"
        )
        self._fill_ai = PatternFill(
            start_color=self.COLOR_AI, 
            end_color=self.COLOR_AI, 
            fill_type="solid"
        )
        self._fill_ao = PatternFill(
            start_color=self.COLOR_AO, 
            end_color=self.COLOR_AO, 
            fill_type="solid"
        )
        self._fill_di = PatternFill(
            start_color=self.COLOR_DI, 
            end_color=self.COLOR_DI, 
            fill_type="solid"
        )
        self._fill_do = PatternFill(
            start_color=self.COLOR_DO, 
            end_color=self.COLOR_DO, 
            fill_type="solid"
        )
        
        self._align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self._align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def generate(self, data: BMSPointsRequest) -> io.BytesIO:
        """
        Gera um Excel estruturado a partir da lista de pontos BMS.
        
        Args:
            data: Dados dos pontos BMS.
            
        Returns:
            Buffer contendo o arquivo Excel gerado.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "BMS Points List"
        
        self._create_header(ws)
        self._populate_data(ws, data)
        self._set_column_widths(ws)
        ws.freeze_panes = "A2"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_header(self, ws):
        """Cria o cabeçalho da planilha."""
        headers = ["Asset Tag", "Point Details", "", "", ""]
        ws.append(headers)
        
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.font = self._font_main_header
            cell.fill = self._fill_main_header
            cell.alignment = self._align_center
            cell.border = self._border_all
        
        ws.merge_cells("B1:E1")
    
    def _populate_data(self, ws, data: BMSPointsRequest):
        """Popula a planilha com os dados agrupados por AssetTag."""
        # Agrupa pontos por AssetTag
        points_by_asset = defaultdict(list)
        for point in data.Points:
            points_by_asset[point.AssetTag].append(point)
        
        current_row = 2
        for asset_tag, points in points_by_asset.items():
            current_row = self._add_asset_group(ws, current_row, asset_tag, points)
    
    def _add_asset_group(self, ws, current_row: int, asset_tag: str, points: list) -> int:
        """Adiciona um grupo de pontos de um asset."""
        num_points = len(points)
        rows_needed = num_points + 1
        end_row = current_row + rows_needed - 1
        
        # Asset Tag
        cell = ws.cell(row=current_row, column=1, value=asset_tag)
        cell.alignment = self._align_top_left
        cell.font = Font(bold=True, size=10)
        
        # Bordas para coluna A
        for r in range(current_row, end_row + 1):
            ws.cell(row=r, column=1).border = self._border_all
        
        # Merge asset tag cell
        if rows_needed > 1:
            ws.merge_cells(
                start_row=current_row, start_column=1, 
                end_row=end_row, end_column=1
            )
        
        # Subcabeçalho
        self._add_subheader(ws, current_row)
        
        # Dados dos pontos
        self._add_points_data(ws, current_row + 1, points)
        
        return current_row + rows_needed
    
    def _add_subheader(self, ws, row: int):
        """Adiciona subcabeçalho dos pontos."""
        sub_headers = ["Point Name", "Point Type", "Logic", "Integration"]
        for i, text in enumerate(sub_headers):
            cell = ws.cell(row=row, column=2+i, value=text)
            cell.font = self._font_sub_header
            cell.fill = self._fill_sub_header
            cell.border = self._border_all
            cell.alignment = self._align_center
    
    def _add_points_data(self, ws, start_row: int, points: list):
        """Adiciona dados dos pontos."""
        for idx, pt in enumerate(points):
            row = start_row + idx
            
            # Point Name
            cell_name = ws.cell(row=row, column=2, value=pt.PointName)
            cell_name.border = self._border_all
            cell_name.alignment = self._align_top_left
            
            # Point Type com cor
            cell_type = ws.cell(row=row, column=3, value=pt.PointType)
            cell_type.border = self._border_all
            cell_type.alignment = self._align_center
            cell_type.fill = self._get_point_type_fill(pt.PointType)
            
            # Logic
            cell_logic = ws.cell(row=row, column=4, value=pt.Logic)
            cell_logic.border = self._border_all
            cell_logic.alignment = self._align_top_left
            
            # Integration
            cell_integration = ws.cell(row=row, column=5, value="Yes" if pt.IsIntegration else "No")
            cell_integration.border = self._border_all
            cell_integration.alignment = self._align_center
            if pt.IsIntegration:
                cell_integration.fill = self._fill_integration
    
    def _get_point_type_fill(self, point_type: str) -> PatternFill:
        """Retorna o fill apropriado para o tipo de ponto."""
        type_fills = {
            "AI": self._fill_ai,
            "AO": self._fill_ao,
            "DI": self._fill_di,
            "DO": self._fill_do,
        }
        
        if point_type in type_fills:
            return type_fills[point_type]
        elif "Integration" in point_type:
            return self._fill_integration
        
        return PatternFill()  # Sem preenchimento
    
    def _set_column_widths(self, ws):
        """Define as larguras das colunas."""
        for i, width in enumerate(self.COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
