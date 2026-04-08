"""
Serviço para geração de Excel no formato Points List (setpoint padrão).

Recebe o JSON gerado pela IA com a estrutura de sistemas/pontos e produz um
arquivo .xlsx com layout idêntico ao "Greymouth Draft Points List MCC.xlsx":

Estrutura do JSON esperado:
{
  "project_name": "...",
  "systems": [
    {
      "system_name":     "...",
      "equipment_tag":   "...",
      "location":        "...",
      "description":     "...",
      "points": [
        {
          "point_description":    "...",
          "types":                ["AI", "HLI"],   // ver mapeamento abaixo
          "field_device_or_notes":"...",
          "qty":                  1
        }
      ]
    }
  ]
}

Mapeamento de tipos JSON → colunas Excel:
  AI        → AI
  XI        → XI
  DI        → DI
  AO        → AO
  XO        → XO
  DO        → DO
  KNX       → KNX
  MOD RTU   → Mod RTU
  MOD IP    → Mod IP
  BAC MS/TP → BAC ms/tp
  HLI       → BAC IP   (High-Level Interface)
  BAC IP    → BAC IP
  PULSE     → M-Bus     (contador de pulsos)
  M-BUS     → M-Bus
"""
import io
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.models import PointsListRequest, PointsListSystem, PointsListPoint


# ──────────────────────────────────────────────────────────────────────────────
# Layout das colunas – espelha o CSV de referência
# ──────────────────────────────────────────────────────────────────────────────
#  Índice (0-based) → nome da coluna no header
#   0  → Description / System Name
#   1  → AI
#   2  → XI
#   3  → DI
#   4  → AO
#   5  → XO
#   6  → DO
#   7  → KNX
#   8  → Mod RTU
#   9  → Mod IP
#  10  → BAC ms/tp
#  11  → BAC IP   (= HLI no JSON)
#  12  → M-Bus    (= Pulse no JSON)
#  13  → Field Device / Notes
#  14  → Qty

COLUMNS: List[Tuple[str, int]] = [
    ("Description",          28),
    ("AI",                    5),
    ("XI",                    5),
    ("DI",                    5),
    ("AO",                    5),
    ("XO",                    5),
    ("DO",                    5),
    ("KNX",                   6),
    ("Mod RTU",               9),
    ("Mod IP",                8),
    ("BAC ms/tp",            11),
    ("BAC IP",                9),
    ("M-Bus",                 8),
    ("Field Device / Notes", 40),
    ("Qty",                   5),
]

# JSON type (uppercase) → índice 0-based da coluna de sinal
TYPE_TO_COL: dict = {
    "AI":        1,
    "XI":        2,
    "DI":        3,
    "AO":        4,
    "XO":        5,
    "DO":        6,
    "KNX":       7,
    "MOD RTU":   8,
    "MOD IP":    9,
    "BAC MS/TP": 10,
    "BAC IP":    11,
    "HLI":       11,   # High-Level Interface → BAC IP
    "M-BUS":     12,
    "PULSE":     12,   # Pulse counter → M-Bus
}


class PointsListExcelService:
    """
    Gera um arquivo Excel (.xlsx) de Points List no formato setpoint padrão.

    O layout replica o "Greymouth Draft Points List MCC.xlsx":
    - Linha 1 : título do projeto (banner)
    - Linha 2 : cabeçalho de colunas com tipos de sinal
    - Por sistema : linha de seção (fundo azul claro) + linhas de pontos
    - Linha final : totais por tipo de sinal
    """

    # ── Cores ─────────────────────────────────────────────────────────────────
    COLOR_TITLE_BG   = "1F4E79"   # azul escuro  – banner do projeto
    COLOR_HEADER_BG  = "2E75B6"   # azul médio   – cabeçalhos de coluna
    COLOR_SECTION_BG = "D6E4F0"   # azul claro   – linha de seção do sistema
    COLOR_AI         = "C6EFCE"   # verde
    COLOR_XI         = "E2EFDA"   # verde claro
    COLOR_AO         = "FFEB9C"   # amarelo
    COLOR_DI         = "BDD7EE"   # azul claro
    COLOR_DO         = "F8CBAD"   # salmão
    COLOR_HLI        = "E2EFDA"   # verde pálido (BAC IP / HLI)
    COLOR_PULSE      = "FCE4D6"   # laranja pálido (M-Bus / Pulse)
    COLOR_TOTAL_BG   = "F2F2F2"   # cinza claro  – linha de totais
    COLOR_WHITE      = "FFFFFF"

    def __init__(self):
        self._setup_styles()

    # ── Configuração de estilos ───────────────────────────────────────────────

    def _setup_styles(self):
        thin = Side(border_style="thin", color="000000")
        self._border = Border(left=thin, right=thin, top=thin, bottom=thin)

        self._font_title   = Font(bold=True, color=self.COLOR_WHITE, size=12)
        self._font_header  = Font(bold=True, color=self.COLOR_WHITE, size=10)
        self._font_section = Font(bold=True, size=9)
        self._font_normal  = Font(size=9)
        self._font_tick    = Font(bold=True, size=9)
        self._font_total   = Font(bold=True, italic=True, size=9)

        self._fill_title   = PatternFill(start_color=self.COLOR_TITLE_BG,   end_color=self.COLOR_TITLE_BG,   fill_type="solid")
        self._fill_header  = PatternFill(start_color=self.COLOR_HEADER_BG,  end_color=self.COLOR_HEADER_BG,  fill_type="solid")
        self._fill_section = PatternFill(start_color=self.COLOR_SECTION_BG, end_color=self.COLOR_SECTION_BG, fill_type="solid")
        self._fill_total   = PatternFill(start_color=self.COLOR_TOTAL_BG,   end_color=self.COLOR_TOTAL_BG,   fill_type="solid")

        # Fill por tipo de sinal (índice 0-based da coluna)
        self._col_fill: dict = {
            1:  PatternFill(start_color=self.COLOR_AI,    end_color=self.COLOR_AI,    fill_type="solid"),
            2:  PatternFill(start_color=self.COLOR_XI,    end_color=self.COLOR_XI,    fill_type="solid"),
            3:  PatternFill(start_color=self.COLOR_DI,    end_color=self.COLOR_DI,    fill_type="solid"),
            4:  PatternFill(start_color=self.COLOR_AO,    end_color=self.COLOR_AO,    fill_type="solid"),
            6:  PatternFill(start_color=self.COLOR_DO,    end_color=self.COLOR_DO,    fill_type="solid"),
            11: PatternFill(start_color=self.COLOR_HLI,   end_color=self.COLOR_HLI,   fill_type="solid"),
            12: PatternFill(start_color=self.COLOR_PULSE,  end_color=self.COLOR_PULSE,  fill_type="solid"),
        }

        self._align_center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self._align_left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        self._align_left_top = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ── Ponto de entrada ──────────────────────────────────────────────────────

    def generate(self, data: PointsListRequest) -> io.BytesIO:
        """
        Gera o arquivo Excel e retorna como buffer BytesIO.

        Args:
            data: Request contendo o nome do projeto e a lista de sistemas.

        Returns:
            BytesIO pronto para ser enviado como StreamingResponse.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Points List"

        self._set_column_widths(ws)
        self._write_title_row(ws, data.project_name or "Unknown")
        self._write_header_row(ws)

        # Acumula totais por coluna de tipo (índice 0-based)
        totals: dict = {i: 0 for i in range(1, 13)}

        current_row = 3
        for system in data.systems:
            current_row = self._write_system_block(ws, current_row, system, totals)

        self._write_totals_row(ws, current_row, totals)

        ws.freeze_panes = "A3"   # congela título + cabeçalho

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Helpers de layout ─────────────────────────────────────────────────────

    def _set_column_widths(self, ws):
        for col_idx, (_, width) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_title_row(self, ws, project_name: str):
        """Linha 1 – banner do projeto mesclado em todas as colunas."""
        num_cols = len(COLUMNS)
        cell = ws.cell(row=1, column=1, value=project_name)
        cell.font      = self._font_title
        cell.fill      = self._fill_title
        cell.alignment = self._align_center
        cell.border    = self._border
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        for c in range(2, num_cols + 1):
            ws.cell(row=1, column=c).border = self._border

    def _write_header_row(self, ws):
        """Linha 2 – cabeçalho de colunas (Description | AI | XI | … | Qty)."""
        for col_idx, (name, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=name)
            cell.font      = self._font_header
            cell.fill      = self._fill_header
            cell.alignment = self._align_center
            cell.border    = self._border

    def _write_system_block(
        self,
        ws,
        start_row: int,
        system: PointsListSystem,
        totals: dict,
    ) -> int:
        """
        Escreve a linha de seção do sistema seguida de todas as linhas de pontos.

        Retorna o próximo número de linha disponível.
        """
        row = start_row
        num_cols = len(COLUMNS)

        # ── Linha de seção ────────────────────────────────────────────────────
        section_label = system.system_name
        if system.equipment_tag:
            section_label = f"{system.system_name}  [{system.equipment_tag}]"

        header_cell = ws.cell(row=row, column=1, value=section_label)
        header_cell.font      = self._font_section
        header_cell.fill      = self._fill_section
        header_cell.alignment = self._align_left
        header_cell.border    = self._border

        # Descrição do sistema na coluna Field Device / Notes
        desc_cell = ws.cell(row=row, column=14, value=system.description or "")
        desc_cell.font      = Font(italic=True, size=9)
        desc_cell.fill      = self._fill_section
        desc_cell.alignment = self._align_left
        desc_cell.border    = self._border

        # Demais células da linha de seção (fundo azul claro, sem valor)
        for c in range(2, num_cols + 1):
            if c != 14:
                cell = ws.cell(row=row, column=c)
                cell.fill   = self._fill_section
                cell.border = self._border

        row += 1

        # ── Linhas de pontos ──────────────────────────────────────────────────
        for point in system.points:
            row = self._write_point_row(ws, row, point, totals)

        # Linha separadora vazia entre sistemas
        for c in range(1, num_cols + 1):
            ws.cell(row=row, column=c).border = self._border
        row += 1

        return row

    def _write_point_row(
        self,
        ws,
        row: int,
        point: PointsListPoint,
        totals: dict,
    ) -> int:
        """
        Escreve uma linha de ponto com ticks nas colunas de tipo correspondentes.

        - Um ponto pode ter múltiplos tipos: cada um recebe um tick "1".
        - Quando qty > 1 o tick exibe qty em vez de 1.

        Retorna a próxima linha disponível.
        """
        qty = point.qty if point.qty is not None else 1

        # Coluna 1 – descrição do ponto
        desc_cell = ws.cell(row=row, column=1, value=point.point_description)
        desc_cell.font      = self._font_normal
        desc_cell.alignment = self._align_left
        desc_cell.border    = self._border

        # Ticks nas colunas de tipo
        ticked_cols: set = set()
        for type_str in point.types:
            col_idx = TYPE_TO_COL.get(type_str.upper())
            if col_idx is not None:
                tick_val = qty if qty > 1 else 1
                cell = ws.cell(row=row, column=col_idx + 1, value=tick_val)
                cell.font      = self._font_tick
                cell.alignment = self._align_center
                cell.border    = self._border
                if col_idx in self._col_fill:
                    cell.fill = self._col_fill[col_idx]
                ticked_cols.add(col_idx)
                totals[col_idx] = totals.get(col_idx, 0) + tick_val

        # Colunas de tipo sem tick (apenas borda)
        for col_idx in range(1, 13):
            if col_idx not in ticked_cols:
                cell = ws.cell(row=row, column=col_idx + 1)
                cell.border    = self._border
                cell.alignment = self._align_center

        # Coluna 14 – notas / field device
        notes_cell = ws.cell(row=row, column=14, value=point.field_device_or_notes or "")
        notes_cell.font      = self._font_normal
        notes_cell.alignment = self._align_left_top
        notes_cell.border    = self._border

        # Coluna 15 – qty
        qty_cell = ws.cell(row=row, column=15, value=qty if qty > 0 else "")
        qty_cell.font      = self._font_normal
        qty_cell.alignment = self._align_center
        qty_cell.border    = self._border

        return row + 1

    def _write_totals_row(self, ws, row: int, totals: dict):
        """Linha de totais – soma de cada coluna de tipo de sinal."""
        num_cols = len(COLUMNS)

        label_cell = ws.cell(row=row, column=1, value="TOTAL")
        label_cell.font      = self._font_total
        label_cell.fill      = self._fill_total
        label_cell.alignment = self._align_left
        label_cell.border    = self._border

        for col_idx in range(1, 13):
            excel_col = col_idx + 1
            val = totals.get(col_idx, 0)
            cell = ws.cell(row=row, column=excel_col, value=val if val else "")
            cell.font      = self._font_total
            cell.alignment = self._align_center
            cell.border    = self._border
            if val and col_idx in self._col_fill:
                cell.fill = self._col_fill[col_idx]
            else:
                cell.fill = self._fill_total

        for c in [14, 15]:
            cell = ws.cell(row=row, column=c)
            cell.fill   = self._fill_total
            cell.border = self._border
